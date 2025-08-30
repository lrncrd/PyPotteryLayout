"""Backend logic refactored and consolidated from notebook helpers.

Provides functions for loading images/metadata, placing images on pages
in grid or puzzle modes, adding captions and scale bars, and saving output.
"""

import os
import re
import random
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import rectpack
import openpyxl

# Default page sizes in pixels (300 DPI approximations)
PAGE_SIZES_PX = {
    'A4': (2480, 3508),
    'A3': (3508, 4961),
    'HD': (1920, 1080),
    '4K': (3840, 2160),
    'LETTER': (2550, 3300),
}


def get_page_dimensions_px(size_name_or_custom, custom_size_str=None):
    if isinstance(size_name_or_custom, tuple) and len(size_name_or_custom) == 2:
        return size_name_or_custom
    if isinstance(size_name_or_custom, str) and 'x' in size_name_or_custom:
        try:
            w, h = map(int, size_name_or_custom.split('x'))
            return (w, h)
        except Exception:
            pass
    if isinstance(size_name_or_custom, str) and size_name_or_custom.lower() == 'custom':
        if not custom_size_str:
            raise ValueError('Specificare custom_size_str se usi "custom"')
        return get_page_dimensions_px(custom_size_str)
    size_px = PAGE_SIZES_PX.get(size_name_or_custom.upper()) if isinstance(size_name_or_custom, str) else None
    if not size_px:
        raise ValueError(f'Formato pagina non supportato: {size_name_or_custom}')
    return size_px


def get_font(size):
    """Try to load common TTF fonts, fallback to default."""
    import platform
    
    # Percorsi comuni per i font su diversi sistemi operativi
    font_paths = []
    
    if platform.system() == "Darwin":  # macOS
        font_paths = [
            "/System/Library/Fonts/Arial.ttf",
            "/System/Library/Fonts/Helvetica.ttc",
            "/System/Library/Fonts/Times.ttc", 
            "/Library/Fonts/Arial.ttf",
            "/Users/" + os.environ.get('USER', 'user') + "/Library/Fonts/Arial.ttf"
        ]
    elif platform.system() == "Windows":
        font_paths = [
            "C:/Windows/Fonts/arial.ttf",
            "C:/Windows/Fonts/calibri.ttf",
            "C:/Windows/Fonts/times.ttf"
        ]
    else:  # Linux
        font_paths = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
            "/usr/share/fonts/TTF/DejaVuSans.ttf"
        ]
    
    # Prova anche i nomi semplici per font di sistema
    simple_names = ["arial.ttf", "Arial.ttf", "DejaVuSans.ttf", "LiberationSans-Regular.ttf", "FreeSans.ttf", "Helvetica.ttf"]
    
    # Prova prima i percorsi completi, poi i nomi semplici
    all_candidates = font_paths + simple_names
    
    for font_path in all_candidates:
        try:
            font = ImageFont.truetype(font_path, int(size))
            # Test che il font sia caricato correttamente testando una dimensione
            test_bbox = font.getbbox("Test")
            if test_bbox and test_bbox[3] > 0:  # Verifica che abbia altezza > 0
                return font
        except Exception:
            continue
    
    # Se nessun TTF funziona, prova il font di default PIL
    try:
        default_font = ImageFont.load_default()
        # Per il font di default, proviamo a creare una versione scalata se possibile
        return default_font
    except Exception:
        # Ultimo fallback
        return ImageFont.load_default()


def get_metadata_headers(filepath):
    """Get column headers from Excel metadata file for GUI dropdown options."""
    if not filepath or not os.path.exists(filepath):
        return None
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        return headers
    except Exception:
        return None


def load_metadata(filepath, status_callback=print):
    if not filepath:
        return None
    status_callback(f"Caricamento metadati da: {filepath}...")
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        metadata = {}
        header = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                metadata[row[0]] = {header[i]: row[i] for i in range(1, len(row))}
        status_callback(f"Caricati metadati per {len(metadata)} elementi.")
        return metadata
    except FileNotFoundError:
        status_callback(f"Attenzione: File metadati '{filepath}' non trovato.")
        return None
    except Exception as e:
        status_callback(f"Errore nel caricamento del file Excel: {e}")
        return None


def load_images_with_info(folder_path, status_callback=print):
    image_data, supported_formats = [], ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')
    status_callback(f"Caricamento immagini da: {folder_path}...")
    if not os.path.isdir(folder_path):
        raise FileNotFoundError(f"'{folder_path}' non esiste.")
    for filename in sorted(os.listdir(folder_path)):
        if filename.lower().endswith(supported_formats):
            try:
                filepath = os.path.join(folder_path, filename)
                img = Image.open(filepath)
                image_data.append({'img': img.copy(), 'name': filename})
                img.close()
            except IOError:
                status_callback(f"Attenzione: Impossibile caricare {filename}.")
    status_callback(f"Caricati {len(image_data)} immagini.")
    return image_data


def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]


def create_scale_bar(target_cm, pixels_per_cm, scale_factor, status_callback=print):
    status_callback(f"Creazione scala grafica per rappresentare {target_cm} cm...")
    try:
        font = get_font(14)
    except Exception:
        font = ImageFont.load_default()
    bar_width_px = int(max(1, target_cm) * pixels_per_cm * scale_factor)
    bar_height_px = 10
    total_height = bar_height_px + 20
    bar_img = Image.new('RGBA', (bar_width_px + 40, total_height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(bar_img)
    num_segments = int(max(1, target_cm))
    segment_width = bar_width_px / num_segments if num_segments else bar_width_px
    for i in range(num_segments):
        color = "black" if i % 2 == 0 else "white"
        x0 = int(i * segment_width) + 20
        x1 = int((i + 1) * segment_width) + 20
        draw.rectangle([x0, 0, x1, bar_height_px], fill=color, outline="black")
    draw.text((20, bar_height_px + 2), "0", fill="black", font=font)
    end_label = f"{target_cm} cm"
    end_label_bbox = draw.textbbox((0, 0), end_label, font=font)
    end_label_width = end_label_bbox[2] - end_label_bbox[0]
    draw.text((20 + bar_width_px - end_label_width, bar_height_px + 2), end_label, fill="black", font=font)
    return bar_img


def scale_images(image_data, scale_factor, status_callback=print):
    if scale_factor == 1.0:
        return image_data
    status_callback(f"Applicazione scala: {scale_factor}x")
    for data in image_data:
        new_width = int(data['img'].width * scale_factor)
        new_height = int(data['img'].height * scale_factor)
        data['img'] = data['img'].resize((new_width, new_height), Image.Resampling.LANCZOS)
    return image_data


def add_captions_to_images(image_data, metadata, font_size, caption_padding, status_callback=print):
    status_callback("Aggiunta didascalie alle immagini...")
    font = get_font(font_size)
    for data in image_data:
        img = data['img']
        caption_lines = [data['name']]
        img_metadata = metadata.get(data['name']) if metadata else None
        if img_metadata:
            for key, value in img_metadata.items():
                if value is not None:
                    caption_lines.append(f"{key}: {value}")
        full_caption_text = "\n".join(caption_lines)
        temp_draw = ImageDraw.Draw(Image.new('RGB', (1, 1)))
        text_bbox = temp_draw.multiline_textbbox((0, 0), full_caption_text, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        new_height = img.height + text_height + caption_padding * 2
        new_width = max(img.width, text_width + caption_padding * 2)
        captioned_img = Image.new('RGB', (new_width, new_height), 'white')
        img_paste_x = (new_width - img.width) // 2
        captioned_img.paste(img, (img_paste_x, 0))
        draw = ImageDraw.Draw(captioned_img)
        text_x = (new_width - text_width) // 2
        text_y = img.height + caption_padding
        draw.multiline_text((text_x, text_y), full_caption_text, font=font, fill="black", align="center")
        data['img'] = captioned_img
    return image_data


def place_images_grid(image_data, page_size_px, grid_size, margin_px, spacing_px, status_callback=print):
    rows_per_page, suggested_cols = grid_size
    page_width, page_height = page_size_px
    available_width = page_width - (2 * margin_px)
    pages, image_index = [], 0
    while image_index < len(image_data):
        current_page = Image.new('RGB', page_size_px, 'white')
        current_y = margin_px
        page_has_images = False
        rows_on_this_page = 0
        while image_index < len(image_data) and rows_on_this_page < rows_per_page:
            row_images, current_row_width, row_height = [], 0, 0
            temp_index = image_index
            while temp_index < len(image_data) and len(row_images) < suggested_cols:
                img = image_data[temp_index]['img']
                needed_width = current_row_width + img.width + (spacing_px if row_images else 0)
                if needed_width <= available_width:
                    row_images.append(image_data[temp_index])
                    current_row_width = needed_width
                    row_height = max(row_height, img.height)
                    temp_index += 1
                else:
                    break
            if not row_images and image_index < len(image_data):
                row_images.append(image_data[image_index])
                row_height = image_data[image_index]['img'].height
                temp_index = image_index + 1
            if not row_images:
                break
            if current_y + row_height > page_height - margin_px:
                image_index = temp_index - len(row_images)
                break
            image_index = temp_index
            total_row_img_width = sum(d['img'].width for d in row_images)
            total_row_width_with_spacing = total_row_img_width + spacing_px * (len(row_images) - 1)
            start_x = margin_px + (available_width - total_row_width_with_spacing) // 2
            current_x = start_x
            for img_data in row_images:
                img = img_data['img']
                paste_y = current_y + (row_height - img.height) // 2
                current_page.paste(img, (current_x, paste_y), img if img.mode == 'RGBA' else None)
                current_x += img.width + spacing_px
                page_has_images = True
            current_y += row_height + spacing_px
            rows_on_this_page += 1
        if page_has_images:
            pages.append(current_page)
        if not page_has_images and image_index < len(image_data):
            status_callback("ATTENZIONE: Immagini rimanenti potrebbero essere troppo grandi.")
            break
    return pages


def place_images_puzzle(image_data, page_size_px, margin_px, spacing_px, status_callback=print):
    page_width, page_height = page_size_px
    bin_width, bin_height = page_width - (2 * margin_px), page_height - (2 * margin_px)
    packer = rectpack.newPacker(rotation=False)
    images = [d['img'] for d in image_data]
    for i, img in enumerate(images):
        packer.add_rect(img.width + spacing_px, img.height + spacing_px, rid=i)
    for _ in range(len(images)):
        packer.add_bin(bin_width, bin_height)
    packer.pack()
    pages = []
    for i, abin in enumerate(packer):
        if not abin:
            break
        page = Image.new('RGB', page_size_px, 'white')
        status_callback(f"Creazione pagina puzzle {i+1}...")
        for rect in abin:
            original_image = images[rect.rid]
            paste_x, paste_y = margin_px + rect.x, margin_px + rect.y
            page.paste(original_image, (paste_x, paste_y), original_image if original_image.mode == 'RGBA' else None)
        pages.append(page)
    return pages


def draw_margin_border(page, margin_px, status_callback=print):
    """Disegna una cornice per visualizzare i margini della pagina."""
    if margin_px <= 0:
        return page
    
    status_callback("Aggiunta cornice margini...")
    draw = ImageDraw.Draw(page)
    
    # Calcola i punti della cornice
    page_width, page_height = page.size
    
    # Rettangolo esterno (bordo pagina)
    outer_rect = [0, 0, page_width - 1, page_height - 1]
    
    # Rettangolo interno (area contenuto)
    inner_rect = [margin_px, margin_px, page_width - margin_px - 1, page_height - margin_px - 1]
    
    # Disegna cornice sottile grigia per mostrare i margini
    # Bordo esterno (pagina)
    draw.rectangle(outer_rect, outline="lightgray", width=1)
    
    # Bordo interno (area contenuto)
    draw.rectangle(inner_rect, outline="gray", width=2)
    
    # Linee d'angolo per enfatizzare i margini
    corner_size = min(20, margin_px // 2)
    if corner_size > 5:
        # Angoli in alto
        draw.line([margin_px, margin_px, margin_px + corner_size, margin_px], fill="darkgray", width=2)
        draw.line([margin_px, margin_px, margin_px, margin_px + corner_size], fill="darkgray", width=2)
        
        draw.line([page_width - margin_px, margin_px, page_width - margin_px - corner_size, margin_px], fill="darkgray", width=2)
        draw.line([page_width - margin_px, margin_px, page_width - margin_px, margin_px + corner_size], fill="darkgray", width=2)
        
        # Angoli in basso
        draw.line([margin_px, page_height - margin_px, margin_px + corner_size, page_height - margin_px], fill="darkgray", width=2)
        draw.line([margin_px, page_height - margin_px, margin_px, page_height - margin_px - corner_size], fill="darkgray", width=2)
        
        draw.line([page_width - margin_px, page_height - margin_px, page_width - margin_px - corner_size, page_height - margin_px], fill="darkgray", width=2)
        draw.line([page_width - margin_px, page_height - margin_px, page_width - margin_px, page_height - margin_px - corner_size], fill="darkgray", width=2)
    
    return page


def save_output(pages, output_file, output_dpi=300, status_callback=print):
    if not pages:
        status_callback("Nessuna pagina generata.")
        return
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if pages and not pages[0].mode == 'RGB':
        pages = [p.convert('RGB') for p in pages]
    status_callback(f"Salvataggio output in: {output_path.resolve()} con metadati a {output_dpi} DPI.")
    if output_path.suffix.lower() == '.pdf':
        pages[0].save(output_path, "PDF", resolution=float(output_dpi), save_all=True, append_images=pages[1:])
    else:
        if len(pages) > 1:
            base, ext = output_path.stem, output_path.suffix
            for i, page in enumerate(pages):
                page.save(output_path.with_name(f"{base}_{i+1}{ext}"), dpi=(output_dpi, output_dpi))
        else:
            pages[0].save(output_path, dpi=(output_dpi, output_dpi))
    status_callback(f"File salvato in: {output_path.resolve()}")


def run_layout_process(params, status_callback=print):
    try:
        metadata = load_metadata(params.get('metadata_file', ''), status_callback)
        image_data = load_images_with_info(params.get('input_folder', ''), status_callback)
        if not image_data:
            status_callback("Nessuna immagine trovata. Processo interrotto.")
            return
        sort_by = params.get('sort_by', '')
        status_callback(f"Ordinamento immagini tramite: '{sort_by or 'alfabetico'}'...")
        if sort_by == 'random':
            random.shuffle(image_data)
        elif sort_by == 'natural_name':
            image_data.sort(key=lambda d: natural_sort_key(d['name']))
        elif sort_by and sort_by not in ['', 'alphabetical', 'random', 'natural_name'] and metadata:
            image_data.sort(key=lambda d: (str(metadata.get(d['name'], {}).get(sort_by, 'zz_fallback')), natural_sort_key(d['name'])))
        image_data = scale_images(image_data, params.get('scale_factor', 1.0), status_callback)
        if params.get('add_caption'):
            image_data = add_captions_to_images(
                image_data,
                metadata,
                params.get('caption_font_size', 14),
                params.get('caption_padding', 4),
                status_callback,
            )
        page_dims = get_page_dimensions_px(params.get('page_size', 'A4'), params.get('custom_size'))
        final_pages = []
        status_callback(f"Avvio posizionamento in modalità '{params.get('mode', 'grid')}'...")
        if params.get('mode') == 'grid':
            grid_size = (params.get('grid_rows', 1), params.get('grid_cols', 1))
            final_pages = place_images_grid(image_data, page_dims, grid_size, params.get('margin_px', 0), params.get('spacing_px', 0), status_callback)
        elif params.get('mode') == 'puzzle':
            final_pages = place_images_puzzle(image_data, page_dims, params.get('margin_px', 0), params.get('spacing_px', 0), status_callback)
        status_callback(f"Generate {len(final_pages)} pagine.")
        if params.get('add_scale_bar') and final_pages:
            if params.get('pixels_per_cm') and params.get('scale_bar_cm'):
                scale_bar = create_scale_bar(params.get('scale_bar_cm', 5), params.get('pixels_per_cm', 100), params.get('scale_factor', 1.0), status_callback)
            else:
                scale_bar = create_scale_bar(params.get('scale_bar_length_px', 100), 1.0, params.get('scale_factor', 1.0), status_callback)
            for page in final_pages:
                x_pos = params.get('margin_px', 0)
                y_pos = page.height - params.get('margin_px', 0) - scale_bar.height
                page.paste(scale_bar, (x_pos, y_pos), scale_bar)
        
        # Aggiungi cornice margini se richiesta
        if params.get('show_margin_border') and final_pages:
            status_callback("Aggiunta cornice margini alle pagine...")
            for i, page in enumerate(final_pages):
                final_pages[i] = draw_margin_border(page, params.get('margin_px', 0), status_callback)
        
        status_callback(f"Salvataggio output in '{params.get('output_file', 'output.pdf')}'...")
        save_output(final_pages, params.get('output_file', 'output.pdf'), params.get('output_dpi', 300), status_callback)
        status_callback('--- PROCESSO COMPLETATO CON SUCCESSO ---')
    except Exception as e:
        status_callback(f"--- ERRORE: {e} ---")
        raise