"""Backend logic refactored with Semantic SVG generation.

Provides functions for loading images/metadata, placing images on pages
in grid or puzzle modes, adding captions and scale bars, and saving output.
Includes semantic SVG export where text and images remain separate/editable.
"""

import os
import re
import random
import io
import base64
from pathlib import Path
from PIL import Image, ImageDraw, ImageFont
import rectpack
import openpyxl

# Default page sizes in pixels (300 DPI approximations)
PAGE_SIZES_PX = {
    'A4': (2480, 3508),
    'A3': (3508, 4961),
    'HD': (1920, 1080),
    '4K': (3840, 2160),
    'LETTER': (2550, 3300),
}

# --- SVG Backend Class ---

class SVGGenerator:
    """Helper class to generate semantic SVG content."""
    def __init__(self, width, height):
        self.width = width
        self.height = height
        self.elements = []
        self.defs = []

    def img_to_base64(self, img):
        """Converts PIL Image to base64 string."""
        buff = io.BytesIO()
        # Convert to RGB if RGBA to save space/avoid issues, unless transparency needed
        save_img = img
        if img.mode != 'RGBA' and img.mode != 'RGB':
            save_img = img.convert('RGB')
        
        save_img.save(buff, format="PNG")
        return base64.b64encode(buff.getvalue()).decode("utf-8")

    def add_image(self, img, x, y, width=None, height=None):
        """Adds an image element."""
        w = width if width else img.width
        h = height if height else img.height
        b64_str = self.img_to_base64(img)
        self.elements.append(
            f'<image x="{x}" y="{y}" width="{w}" height="{h}" href="data:image/png;base64,{b64_str}"/>'
        )

    def add_text(self, text, x, y, font_size, font_family="Arial", anchor="start", color="black"):
        """Adds a text element. Handles multi-line text via tspan."""
        lines = text.split('\n')
        line_height = font_size * 1.2
        
        text_xml = f'<text x="{x}" y="{y}" font-family="{font_family}" font-size="{font_size}" fill="{color}" text-anchor="{anchor}">'
        
        # For the first line, we use the y passed. For subsequent, we use dy.
        for i, line in enumerate(lines):
            # XML escape for safety
            safe_line = line.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            dy = 0 if i == 0 else line_height
            # If anchor is middle, x must be maintained for tspans
            text_xml += f'<tspan x="{x}" dy="{dy}">{safe_line}</tspan>'
        
        text_xml += '</text>'
        self.elements.append(text_xml)

    def add_rect(self, x, y, width, height, fill="none", stroke="none", stroke_width=0):
        """Adds a rectangle (used for dividers, scale bars, etc)."""
        self.elements.append(
            f'<rect x="{x}" y="{y}" width="{width}" height="{height}" fill="{fill}" stroke="{stroke}" stroke-width="{stroke_width}"/>'
        )

    def add_line(self, x1, y1, x2, y2, stroke="black", stroke_width=1):
        """Adds a line element."""
        self.elements.append(
            f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="{stroke}" stroke-width="{stroke_width}"/>'
        )

    def get_xml(self):
        """Returns the full SVG XML string."""
        header = f'<svg width="{self.width}" height="{self.height}" viewBox="0 0 {self.width} {self.height}" xmlns="http://www.w3.org/2000/svg" version="1.1">'
        body = "\n".join(self.elements)
        footer = '</svg>'
        return header + "\n" + body + "\n" + footer


# --- Core Functions ---

def get_page_dimensions_px(size_name_or_custom, custom_size_str=None):
    if isinstance(size_name_or_custom, tuple) and len(size_name_or_custom) == 2:
        return size_name_or_custom
    if isinstance(size_name_or_custom, str) and 'x' in size_name_or_custom:
        try:
            w, h = map(int, size_name_or_custom.split('x'))
            return (w, h)
        except Exception:
            pass
    if isinstance(size_name_or_custom, str) and size_name_or_custom.lower() == 'custom':
        if not custom_size_str:
            raise ValueError('Specificare custom_size_str se usi "custom"')
        return get_page_dimensions_px(custom_size_str)
    size_px = PAGE_SIZES_PX.get(size_name_or_custom.upper()) if isinstance(size_name_or_custom, str) else None
    if not size_px:
        raise ValueError(f'Unsupported page format: {size_name_or_custom}')
    return size_px


def get_font(size):
    """Try to load common TTF fonts, fallback to default."""
    import platform
    font_paths = []
    if platform.system() == "Darwin":  # macOS
        font_paths = ["/Library/Fonts/Arial.ttf", "/System/Library/Fonts/Helvetica.ttc"]
    elif platform.system() == "Windows":
        font_paths = ["C:/Windows/Fonts/arial.ttf", "C:/Windows/Fonts/calibri.ttf"]
    else:  # Linux
        font_paths = ["/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
    
    simple_names = ["arial.ttf", "Arial.ttf", "DejaVuSans.ttf"]
    all_candidates = font_paths + simple_names
    
    for font_path in all_candidates:
        try:
            font = ImageFont.truetype(font_path, int(size))
            if font.getbbox("Test")[3] > 0: return font
        except Exception: continue
    
    return ImageFont.load_default()


def get_metadata_headers(filepath):
    if not filepath or not os.path.exists(filepath): return None
    try:
        workbook = openpyxl.load_workbook(filepath)
        return [cell.value for cell in workbook.active[1]]
    except Exception: return None


def load_metadata(filepath, status_callback=print):
    if not filepath: return None
    status_callback(f"Loading metadata from: {filepath}...")
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
        metadata = {}
        header = [cell.value for cell in sheet[1]]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                metadata[row[0]] = {header[i]: row[i] for i in range(1, len(row))}
        status_callback(f"Loaded metadata for {len(metadata)} items.")
        return metadata
    except Exception as e:
        status_callback(f"Error loading Excel file: {e}")
        return None


def load_images_with_info(folder_path, status_callback=print):
    image_data = []
    supported_formats = ('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')
    status_callback(f"Loading images from: {folder_path}...")
    if not os.path.isdir(folder_path):
        raise FileNotFoundError(f"'{folder_path}' does not exist.")
    
    for filename in sorted(os.listdir(folder_path)):
        if filename.lower().endswith(supported_formats):
            try:
                filepath = os.path.join(folder_path, filename)
                img = Image.open(filepath)
                # Store simple dict initially. SVG components added later.
                image_data.append({'img': img.copy(), 'name': filename})
                img.close()
            except IOError:
                status_callback(f"Warning: Could not load {filename}.")
    
    status_callback(f"Loaded {len(image_data)} images.")
    return image_data


def natural_sort_key(s):
    return [int(text) if text.isdigit() else text.lower() for text in re.split(r'(\d+)', s)]


def sort_images_hierarchical(image_data, primary_sort, secondary_sort, metadata, status_callback=print):
    if not image_data: return image_data
    if not primary_sort or primary_sort in ['', 'alphabetical']: primary_sort = 'alphabetical'
    
    status_callback(f"Sorting: '{primary_sort}' -> '{secondary_sort}'...")
    
    def get_sort_key(img_data, sort_field):
        if sort_field == 'random': return (0, random.random(), '')
        elif sort_field == 'natural_name': return (2, 0, natural_sort_key(img_data['name']))
        elif sort_field == 'alphabetical': return (2, 0, img_data['name'].lower())
        else:
            if metadata and img_data['name'] in metadata:
                value = metadata[img_data['name']].get(sort_field, '')
                if value is None: return (2, 0, 'zzz_empty')
                try: return (1, float(str(value).strip()), '')
                except ValueError: return (2, 0, str(value).lower())
            return (2, 0, 'zzz_no_metadata')

    if primary_sort == 'random' and (not secondary_sort or secondary_sort == 'none'):
        random.shuffle(image_data)
    else:
        def composite_sort_key(img_data):
            p_key = get_sort_key(img_data, primary_sort)
            s_key = get_sort_key(img_data, secondary_sort) if secondary_sort and secondary_sort != 'none' else (0,0,'')
            return (p_key, s_key, (2, 0, natural_sort_key(img_data['name'])))
        image_data.sort(key=composite_sort_key)
    
    return image_data


def create_scale_bar(target_cm, pixels_per_cm, scale_factor, status_callback=print):
    """
    Creates a scale bar.
    Returns a tuple: (PIL_Image, SVG_Instructions_Dict)
    """
    status_callback(f"Creating scale bar representing {target_cm} cm...")
    
    try: font = get_font(14)
    except: font = ImageFont.load_default()
    
    bar_width_px = int(max(1, target_cm) * pixels_per_cm * scale_factor)
    bar_height_px = 10
    total_height = bar_height_px + 20
    bar_total_width = bar_width_px + 40
    
    # 1. Create Raster Image (PIL)
    bar_img = Image.new('RGBA', (bar_total_width, total_height), (0, 0, 0, 0))
    draw = ImageDraw.Draw(bar_img)
    
    num_segments = int(max(1, target_cm))
    segment_width = bar_width_px / num_segments if num_segments else bar_width_px
    
    # Store semantic data for SVG
    svg_data = {
        'type': 'scale_bar',
        'width': bar_total_width,
        'height': total_height,
        'segments': [],
        'label': f"{target_cm} cm",
        'font_size': 14
    }

    for i in range(num_segments):
        color = "black" if i % 2 == 0 else "white"
        x0 = int(i * segment_width) + 20
        x1 = int((i + 1) * segment_width) + 20
        
        # Draw raster
        draw.rectangle([x0, 0, x1, bar_height_px], fill=color, outline="black")
        
        # Store vector info
        svg_data['segments'].append({
            'x': x0, 'y': 0, 'w': x1-x0, 'h': bar_height_px, 'fill': color
        })

    # Labels
    draw.text((20, bar_height_px + 2), "0", fill="black", font=font)
    
    end_label = f"{target_cm} cm"
    end_label_bbox = draw.textbbox((0, 0), end_label, font=font)
    end_label_width = end_label_bbox[2] - end_label_bbox[0]
    end_label_x = 20 + bar_width_px - end_label_width
    draw.text((end_label_x, bar_height_px + 2), end_label, fill="black", font=font)

    return bar_img, svg_data


def scale_images(image_data, scale_factor, status_callback=print):
    if scale_factor == 1.0: return image_data
    status_callback(f"Applying scale: {scale_factor}x")
    for data in image_data:
        # If we have semantic SVG components, we need to scale their layout info?
        # For simplicity, we scale the base image. SVG components are generated relative to this.
        new_width = int(data['img'].width * scale_factor)
        new_height = int(data['img'].height * scale_factor)
        data['img'] = data['img'].resize((new_width, new_height), Image.Resampling.LANCZOS)
    return image_data


def add_captions_to_images(image_data, metadata, font_size, caption_padding, remove_extension=False, selected_fields=None, hide_field_names=False, status_callback=print):
    """
    Adds captions. 
    CRITICAL UPDATE: This function now prepares data for BOTH raster (baked pixels) and SVG (semantic text).
    """
    status_callback("Adding captions to images...")
    font = get_font(font_size)
    
    for data in image_data:
        original_img = data['img']
        
        # 1. Prepare Text
        filename = data['name']
        if remove_extension: filename = os.path.splitext(filename)[0]
        caption_lines = [filename]
        
        img_metadata = metadata.get(data['name']) if metadata else None
        if img_metadata:
            fields_to_use = selected_fields if selected_fields else img_metadata.keys()
            for key in fields_to_use:
                val = img_metadata.get(key)
                if val is not None:
                    caption_lines.append(str(val) if hide_field_names else f"{key}: {val}")
        
        full_caption_text = "\n".join(caption_lines)
        
        # 2. Calculate Dimensions
        temp_draw = ImageDraw.Draw(Image.new('RGB', (1, 1)))
        text_bbox = temp_draw.multiline_textbbox((0, 0), full_caption_text, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        
        new_height = original_img.height + text_height + caption_padding * 2
        new_width = max(original_img.width, text_width + caption_padding * 2)
        
        # 3. Create Raster Image (Preview/JPG output)
        captioned_img = Image.new('RGB', (new_width, new_height), 'white')
        img_paste_x = (new_width - original_img.width) // 2
        captioned_img.paste(original_img, (img_paste_x, 0))
        
        draw = ImageDraw.Draw(captioned_img)
        text_x = (new_width - text_width) // 2
        text_y = original_img.height + caption_padding
        draw.multiline_text((text_x, text_y), full_caption_text, font=font, fill="black", align="center")
        
        # 4. Update Data Structure
        # 'img' is now the baked version for layout engine and raster export
        data['img'] = captioned_img
        
        # 'svg_components' stores separated semantic data for SVG export
        data['svg_components'] = {
            'type': 'captioned_image',
            'original_img': original_img, # Keep reference to clean image
            'caption_text': full_caption_text,
            'font_size': font_size,
            'dims': {
                # Relative coordinates within the bounding box
                'img_x': img_paste_x,
                'img_y': 0,
                'text_x': new_width // 2, # Center for text anchor
                'text_y': text_y
            }
        }
        
    return image_data


def _render_item_to_svg(svg_gen, item_data, abs_x, abs_y):
    """Helper to render a generic item (image, captioned image, or scale bar) to the SVG Generator."""
    
    # Case A: Semantic Captioned Image (Created by add_captions_to_images)
    if 'svg_components' in item_data and item_data['svg_components']['type'] == 'captioned_image':
        comp = item_data['svg_components']
        dims = comp['dims']
        
        # 1. Place the clean photo
        svg_gen.add_image(
            comp['original_img'], 
            x = abs_x + dims['img_x'], 
            y = abs_y + dims['img_y']
        )
        
        # 2. Place the editable text
        # Note: SVG text y is the baseline. PIL text y is top-left. 
        # We approximate baseline shift or just use the y provided.
        svg_gen.add_text(
            comp['caption_text'],
            x = abs_x + dims['text_x'],
            y = abs_y + dims['text_y'] + comp['font_size'], # Shift down slightly for baseline
            font_size = comp['font_size'],
            anchor = "middle"
        )
        
    # Case B: Semantic Scale Bar
    elif 'svg_components' in item_data and item_data['svg_components']['type'] == 'scale_bar':
        svg_data = item_data['svg_components']
        
        # Draw segments
        for seg in svg_data['segments']:
            svg_gen.add_rect(
                x = abs_x + seg['x'],
                y = abs_y + seg['y'],
                width = seg['w'],
                height = seg['h'],
                fill = seg['fill'],
                stroke = "black",
                stroke_width = 1
            )
        
        # Draw text labels (0 and Total)
        # Label 0
        svg_gen.add_text("0", abs_x + 20, abs_y + svg_data['height'] - 5, svg_data['font_size'])
        
        # Label End
        svg_gen.add_text(
            svg_data['label'], 
            abs_x + svg_data['width'] - 20, # This needs anchor end or precise calculation
            abs_y + svg_data['height'] - 5, 
            svg_data['font_size'], 
            anchor="end"
        )

    # Case C: Generic Image (No captions, just a bitmap)
    else:
        svg_gen.add_image(item_data['img'], abs_x, abs_y)


def place_images_grid(image_data, page_size_px, grid_size, margin_px, spacing_px, 
                      page_break_on_primary_change=False, primary_sort_key=None, 
                      primary_break_type='new_page', divider_thickness=5, divider_width_percent=80,
                      vertical_alignment='center', status_callback=print):
    
    rows_per_page, suggested_cols = grid_size
    page_width, page_height = page_size_px
    available_width = page_width - (2 * margin_px)
    available_height = page_height - (2 * margin_px)
    
    # Output lists
    pil_pages = []
    svg_pages = [] # List of SVG strings
    
    image_index = 0
    total_images = len(image_data)
    current_primary_value = None
    
    status_callback(f"Starting grid layout: {total_images} images to place")
    
    while image_index < len(image_data):
        # Initialize PIL Page
        current_pil_page = Image.new('RGB', page_size_px, 'white')
        # Initialize SVG Generator
        current_svg_gen = SVGGenerator(page_width, page_height)
        # Add white background rect for SVG
        current_svg_gen.add_rect(0, 0, page_width, page_height, fill="white")
        
        page_has_images = False
        
        if page_break_on_primary_change and primary_sort_key and image_index < len(image_data):
            page_primary_value = primary_sort_key(image_data[image_index])
            if current_primary_value is not None and page_primary_value != current_primary_value:
                status_callback(f"Page break: primary sort changed from {current_primary_value} to {page_primary_value}")
            current_primary_value = page_primary_value
        
        page_rows = []
        divider_rows = []
        temp_image_index = image_index
        temp_rows_on_page = 0
        
        # Layout Calculation (Identify rows)
        while temp_image_index < len(image_data) and temp_rows_on_page < rows_per_page:
            row_images, current_row_width, row_height = [], 0, 0
            temp_index = temp_image_index
            
            if page_break_on_primary_change and primary_sort_key and temp_rows_on_page > 0:
                if temp_image_index < len(image_data):
                    next_primary_value = primary_sort_key(image_data[temp_image_index])
                    if next_primary_value != current_primary_value:
                        if primary_break_type == 'new_page': break
                        else:
                            divider_rows.append((temp_rows_on_page, next_primary_value))
                            current_primary_value = next_primary_value
            
            while temp_index < len(image_data) and len(row_images) < suggested_cols:
                if page_break_on_primary_change and primary_sort_key and len(row_images) > 0:
                    if temp_index < len(image_data):
                        if primary_sort_key(image_data[temp_index]) != current_primary_value: break
                
                img = image_data[temp_index]['img']
                needed_width = current_row_width + img.width + (spacing_px if row_images else 0)
                if needed_width <= available_width:
                    row_images.append(image_data[temp_index])
                    current_row_width = needed_width
                    row_height = max(row_height, img.height)
                    temp_index += 1
                else: break
            
            if not row_images and temp_image_index < len(image_data):
                row_images.append(image_data[temp_image_index])
                row_height = image_data[temp_image_index]['img'].height
                temp_index = temp_image_index + 1
            
            if not row_images: break
            page_rows.append((row_images, row_height))
            temp_image_index = temp_index
            temp_rows_on_page += 1
        
        if not page_rows: break
        
        # Calculate Vertical Spacing
        divider_margin = 20
        total_content_height = sum(r[1] for r in page_rows)
        total_spacing_height = spacing_px * (len(page_rows) - 1) if len(page_rows) > 1 else 0
        total_separator_height = len(divider_rows) * (divider_thickness + 2 * divider_margin)
        total_height_needed = total_content_height + total_spacing_height + total_separator_height
        
        start_y = margin_px
        if vertical_alignment == 'center' and total_height_needed < available_height:
            start_y = margin_px + (available_height - total_height_needed) // 2
        
        # Rendering (Both PIL and SVG)
        current_y = start_y
        images_placed_on_page = 0
        divider_dict = {row_idx: val for row_idx, val in divider_rows}
        
        for row_idx, (row_images, row_height) in enumerate(page_rows):
            # Render Divider
            if row_idx in divider_dict:
                # PIL Drawing
                pil_draw = ImageDraw.Draw(current_pil_page)
                divider_y = current_y + divider_margin
                divider_width = int(available_width * (divider_width_percent / 100))
                div_start_x = margin_px + (available_width - divider_width) // 2
                div_end_x = div_start_x + divider_width
                pil_draw.line([(div_start_x, divider_y), (div_end_x, divider_y)], fill='black', width=divider_thickness)
                
                # SVG Drawing (Semantic Line)
                current_svg_gen.add_line(div_start_x, divider_y, div_end_x, divider_y, stroke="black", stroke_width=divider_thickness)
                
                current_y += divider_thickness + 2 * divider_margin

            if current_y + row_height > page_height - margin_px: break
                
            # Render Row Images
            total_row_img_width = sum(d['img'].width for d in row_images)
            total_row_width_with_spacing = total_row_img_width + spacing_px * (len(row_images) - 1)
            current_x = margin_px + (available_width - total_row_width_with_spacing) // 2
            
            for img_data in row_images:
                img = img_data['img']
                paste_y = current_y + (row_height - img.height) // 2
                
                # 1. PIL Paste
                current_pil_page.paste(img, (current_x, paste_y), img if img.mode == 'RGBA' else None)
                
                # 2. SVG Add (Semantic)
                _render_item_to_svg(current_svg_gen, img_data, current_x, paste_y)
                
                current_x += img.width + spacing_px
                page_has_images = True
                images_placed_on_page += 1
            
            current_y += row_height + spacing_px
            image_index += len(row_images)
        
        if page_has_images:
            pil_pages.append(current_pil_page)
            svg_pages.append(current_svg_gen.get_xml())
            status_callback(f"Page {len(pil_pages)} created with {images_placed_on_page} images")
    
    # Handle Leftovers (Simplified logic for brevity, same parallel approach applies)
    if image_index < total_images:
        remaining = image_data[image_index:]
        for img_data in remaining:
            # Create single page
            p = Image.new('RGB', page_size_px, 'white')
            s = SVGGenerator(page_width, page_height)
            s.add_rect(0, 0, page_width, page_height, fill="white")
            
            img = img_data['img']
            # Scale logic (omitted for brevity, assume fits or scaled previously)
            px = (page_width - img.width) // 2
            py = (page_height - img.height) // 2
            
            p.paste(img, (px, py))
            _render_item_to_svg(s, img_data, px, py)
            
            pil_pages.append(p)
            svg_pages.append(s.get_xml())
            status_callback(f"Created individual page for leftover: {img_data.get('name')}")

    return pil_pages, svg_pages


def place_images_puzzle(image_data, page_size_px, margin_px, spacing_px, 
                        page_break_on_primary_change=False, primary_sort_key=None, status_callback=print):
    
    # Wrapper to handle grouping, then delegates to internal
    if page_break_on_primary_change and primary_sort_key:
        from collections import OrderedDict
        groups = OrderedDict()
        for d in image_data:
            k = primary_sort_key(d)
            if k not in groups: groups[k] = []
            groups[k].append(d)
        
        all_pil, all_svg = [], []
        for k, g_imgs in groups.items():
            p, s = _place_images_puzzle_internal(g_imgs, page_size_px, margin_px, spacing_px, status_callback)
            all_pil.extend(p)
            all_svg.extend(s)
        return all_pil, all_svg
    else:
        return _place_images_puzzle_internal(image_data, page_size_px, margin_px, spacing_px, status_callback)


def _place_images_puzzle_internal(image_data, page_size_px, margin_px, spacing_px, status_callback=print):
    page_width, page_height = page_size_px
    bin_width = page_width - (2 * margin_px)
    bin_height = page_height - (2 * margin_px)
    
    packer = rectpack.newPacker(rotation=False)
    images = [d['img'] for d in image_data]
    
    # Map rectpack ID back to image_data index
    for i, img in enumerate(images):
        packer.add_rect(img.width + spacing_px, img.height + spacing_px, rid=i)
    
    # Add enough bins
    for _ in range(len(images)):
        packer.add_bin(bin_width, bin_height)
    
    packer.pack()
    
    pil_pages = []
    svg_pages = []
    placed_indices = set()
    
    for i, abin in enumerate(packer):
        if not abin: break
        
        current_pil = Image.new('RGB', page_size_px, 'white')
        current_svg = SVGGenerator(page_width, page_height)
        current_svg.add_rect(0, 0, page_width, page_height, fill="white")
        
        count = 0
        for rect in abin:
            img_idx = rect.rid
            placed_indices.add(img_idx)
            data = image_data[img_idx]
            img = data['img']
            
            x = margin_px + rect.x
            y = margin_px + rect.y
            
            # 1. PIL
            current_pil.paste(img, (x, y), img if img.mode == 'RGBA' else None)
            
            # 2. SVG
            _render_item_to_svg(current_svg, data, x, y)
            count += 1
            
        pil_pages.append(current_pil)
        svg_pages.append(current_svg.get_xml())
        status_callback(f"Puzzle Page {i+1}: {count} images")
    
    # Handle unplaced (single pages)
    remaining_indices = set(range(len(image_data))) - placed_indices
    for idx in remaining_indices:
        d = image_data[idx]
        img = d['img']
        
        p = Image.new('RGB', page_size_px, 'white')
        s = SVGGenerator(page_width, page_height)
        s.add_rect(0, 0, page_width, page_height, fill="white")
        
        # Simple center logic
        # (Scaling logic omitted for brevity, assume pre-scaled or fits)
        x = (page_width - img.width) // 2
        y = (page_height - img.height) // 2
        
        p.paste(img, (x, y))
        _render_item_to_svg(s, d, x, y)
        
        pil_pages.append(p)
        svg_pages.append(s.get_xml())
        status_callback(f"Individual page for unplaced puzzle image: {d['name']}")
        
    return pil_pages, svg_pages